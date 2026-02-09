"""
Import existing fishing organizations from Excel file into SQLite database.
Usage: python import_excel.py path/to/fishing_organizations_database.xlsx
"""

import sys
import sqlite3
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

DATABASE_PATH = "fishing_organizations.db"

# Mapping of sheet names to organization types
SHEET_TYPE_MAP = {
    "Conservation-Advocacy": "Conservation",
    "Fishing-Clubs": "Fishing Club",
    "Tournaments": "Tournament",
    "Charter-Guide Assoc": "Guide Association",
    "Fly Fishing": "Fly Fishing",
    "Media-Publications": "Media/Publication",
    "South-Atlantic": "Regional Organization",
    "Gulf-Coast": "Regional Organization",
    "Mid-Atlantic": "Regional Organization",
    "Northeast": "Regional Organization",
}

def init_database():
    """Initialize the SQLite database."""
    conn = sqlite3.connect(DATABASE_PATH)
    cursor = conn.cursor()
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS organizations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            org_type TEXT,
            focus_area TEXT,
            state_region TEXT,
            website TEXT,
            contact TEXT,
            membership TEXT,
            description TEXT,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_state ON organizations(state_region)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_type ON organizations(org_type)")
    
    conn.commit()
    return conn

def get_header_mapping(headers):
    """Map Excel column headers to database fields."""
    mapping = {}
    header_aliases = {
        'name': ['name', 'organization', 'org name', 'organization name', 'club name'],
        'org_type': ['type', 'org type', 'organization type', 'category'],
        'focus_area': ['focus', 'focus area', 'specialty', 'fishing type'],
        'state_region': ['state', 'region', 'state/region', 'location', 'area'],
        'website': ['website', 'url', 'web', 'site'],
        'contact': ['contact', 'phone', 'email', 'contact info'],
        'membership': ['membership', 'members', 'size'],
        'description': ['description', 'about', 'details', 'info'],
        'notes': ['notes', 'comments', 'other'],
    }
    
    for idx, header in enumerate(headers):
        if header is None:
            continue
        header_lower = str(header).lower().strip()
        for field, aliases in header_aliases.items():
            if header_lower in aliases:
                mapping[idx] = field
                break
    
    return mapping

def import_sheet(conn, sheet, sheet_name):
    """Import a single sheet into the database."""
    cursor = conn.cursor()
    rows = list(sheet.iter_rows(values_only=True))
    
    if len(rows) < 2:
        print(f"  Skipping {sheet_name}: no data rows")
        return 0
    
    headers = rows[0]
    mapping = get_header_mapping(headers)
    
    if 'name' not in [mapping.get(i) for i in range(len(headers))]:
        # Try to find name column by checking first non-empty column
        for idx, header in enumerate(headers):
            if header and str(header).strip():
                mapping[idx] = 'name'
                break
    
    # Determine default org_type from sheet name
    default_type = SHEET_TYPE_MAP.get(sheet_name, sheet_name)
    
    imported = 0
    for row in rows[1:]:
        if not row or not any(row):
            continue
        
        org_data = {
            'name': None,
            'org_type': default_type,
            'focus_area': None,
            'state_region': None,
            'website': None,
            'contact': None,
            'membership': None,
            'description': None,
            'notes': None,
        }
        
        for idx, value in enumerate(row):
            if idx in mapping and value is not None:
                field = mapping[idx]
                org_data[field] = str(value).strip() if value else None
        
        # Skip if no name
        if not org_data['name']:
            continue
        
        # Check for duplicates
        cursor.execute("SELECT id FROM organizations WHERE name = ?", (org_data['name'],))
        if cursor.fetchone():
            continue
        
        cursor.execute("""
            INSERT INTO organizations (name, org_type, focus_area, state_region, website, contact, membership, description, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            org_data['name'],
            org_data['org_type'],
            org_data['focus_area'],
            org_data['state_region'],
            org_data['website'],
            org_data['contact'],
            org_data['membership'],
            org_data['description'],
            org_data['notes'],
        ))
        imported += 1
    
    conn.commit()
    return imported

def main():
    if len(sys.argv) < 2:
        print("Usage: python import_excel.py <excel_file.xlsx>")
        print("\nThis script imports fishing organizations from an Excel file into the SQLite database.")
        sys.exit(1)
    
    excel_path = Path(sys.argv[1])
    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    print(f"Loading Excel file: {excel_path}")
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    
    print(f"Found {len(workbook.sheetnames)} sheets: {', '.join(workbook.sheetnames)}")
    
    conn = init_database()
    total_imported = 0
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        imported = import_sheet(conn, sheet, sheet_name)
        print(f"  {sheet_name}: imported {imported} organizations")
        total_imported += imported
    
    conn.close()
    workbook.close()
    
    print(f"\nTotal imported: {total_imported} organizations")
    print(f"Database saved to: {DATABASE_PATH}")

if __name__ == "__main__":
    main()
